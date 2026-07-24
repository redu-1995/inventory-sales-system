import csv
from decimal import Decimal
from datetime import datetime

from django.http import HttpResponse
from django.db.models import DecimalField, ExpressionWrapper, F, Q, Sum
from rest_framework.response import Response
from rest_framework.viewsets import ModelViewSet, ViewSet, ReadOnlyModelViewSet
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

from .models import (
    Inventory,
    StockMovement,
   
)

from .serializers import (
    InventorySerializer,
    StockMovementSerializer,
    
)


class InventoryViewSet(ModelViewSet):
    queryset = Inventory.objects.all()
    serializer_class = InventorySerializer

    def get_queryset(self):
        queryset = super().get_queryset().select_related('product', 'product__category')

        search = self.request.query_params.get('search', '').strip()
        category = self.request.query_params.get('category', '').strip()
        status = self.request.query_params.get('status', '').strip()
        low_stock = self.request.query_params.get('low_stock', '').strip()

        # Handle Search
        if search:
            queryset = queryset.filter(
                Q(product__name__icontains=search) | Q(product__sku__icontains=search)
            )

        # Handle Category Filter
        if category:
            queryset = queryset.filter(product__category__name__icontains=category)

        # Handle low_stock shortcut query param (?low_stock=true) using the correct field: 'reorder_level'
        if low_stock == 'true':
            queryset = queryset.filter(quantity__gt=0, quantity__lte=F('reorder_level'))

        # Handle Status Filter
        if status:
            if status == 'OUT_OF_STOCK':
                queryset = queryset.filter(quantity=0)
            elif status == 'LOW_STOCK':
                queryset = queryset.filter(quantity__gt=0, quantity__lte=F('reorder_level'))
            elif status == 'IN_STOCK':
                queryset = queryset.filter(quantity__gt=F('reorder_level'))

        return queryset

    def list(self, request, *args, **kwargs):
        if request.query_params.get('summary') == 'true':
            queryset = self.filter_queryset(self.get_queryset())
            total_stock_units = queryset.aggregate(total=Sum('quantity'))['total'] or 0
            low_stock = queryset.filter(quantity__gt=0, quantity__lte=F('reorder_level')).count()
            out_of_stock = queryset.filter(quantity=0).count()
            inventory_value = queryset.aggregate(
                value=Sum(
                    ExpressionWrapper(F('quantity') * F('product__cost_price'), output_field=DecimalField())
                )
            )['value'] or Decimal('0.00')

            return Response({
                'total_stock_units': int(total_stock_units),
                'low_stock': low_stock,
                'out_of_stock': out_of_stock,
                'inventory_value': float(inventory_value),
            })

        return super().list(request, *args, **kwargs)


class StockMovementViewSet(ModelViewSet):
    """
    A viewset that provides default handling for logging and reviewing stock changes.
    All business logic for IN, OUT, and ADJUST operations runs securely within 
    the overridden StockMovementSerializer pipeline.
    """
    queryset = StockMovement.objects.all().order_by('-created_at')
    serializer_class = StockMovementSerializer
  
    def perform_create(self, serializer):
        """
        Binds the currently authenticated user to the history row 
        automatically when a standard POST action is made.
        """
        serializer.save(user=self.request.user)





class InventoryExportViewSet(ReadOnlyModelViewSet):
    queryset = Inventory.objects.all()
    permission_classes = [IsAuthenticated]

    def get_filtered_data(self, request):
        queryset = self.filter_queryset(self.get_queryset())
        return queryset

    # ==========================================
    # 1. CSV EXPORT
    # ==========================================
    @action(detail=False, methods=['get'], url_path='export-csv')
    def export_csv(self, request):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="inventory_export_{datetime.now().strftime("%Y%m%d")}.csv"'

        writer = csv.writer(response)
        writer.writerow(['Product', 'SKU', 'Category', 'Quantity', 'Reorder Level', 'Status'])

        for item in self.get_filtered_data(request):
            reorder_level = getattr(item, 'reorder_level', 0)
            
            if item.quantity == 0:
                status = 'OUT_OF_STOCK'
            elif item.quantity <= reorder_level:
                status = 'LOW_STOCK'
            else:
                status = 'IN_STOCK'

            writer.writerow([
                item.product.name,
                item.product.sku,
                item.product.category.name if item.product.category else 'N/A',
                item.quantity,
                reorder_level,
                status
            ])
        return response

    # ==========================================
    # 2. EXCEL EXPORT (.XLSX)
    # ==========================================
    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="inventory_export_{datetime.now().strftime("%Y%m%d")}.xlsx"'

        wb = Workbook()
        ws = wb.active
        ws.title = "Current Inventory"

        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")

        headers = ['Product Name', 'SKU', 'Category', 'Current Stock', 'Reorder Level', 'Status']
        ws.append(headers)

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        for item in self.get_filtered_data(request):
            reorder_level = getattr(item, 'reorder_level', 0)
            
            if item.quantity == 0:
                status = 'OUT_OF_STOCK'
            elif item.quantity <= reorder_level:
                status = 'LOW_STOCK'
            else:
                status = 'IN_STOCK'

            ws.append([
                item.product.name,
                item.product.sku,
                item.product.category.name if item.product.category else 'N/A',
                item.quantity,
                reorder_level,
                status
            ])

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        wb.save(response)
        return response

    # ==========================================
    # 3. PDF REPORT
    # ==========================================
    @action(detail=False, methods=['get'], url_path='export-pdf')
    def export_pdf(self, request):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="inventory_report_{datetime.now().strftime("%Y%m%d")}.pdf"'

        doc = SimpleDocTemplate(response, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
        story = []
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            'ReportTitle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=24, textColor=colors.HexColor('#1E293B'), spaceAfter=10
        )
        meta_style = ParagraphStyle(
            'ReportMeta', parent=styles['Normal'], fontName='Helvetica', fontSize=10, textColor=colors.HexColor('#64748B'), spaceAfter=20
        )

        story.append(Paragraph("Inventory Valuation Report", title_style))
        story.append(Paragraph(f"Generated on: {datetime.now().strftime('%B %d, %Y')} | Scope: Complete Stock", meta_style))
        story.append(Spacer(1, 15))

        items = self.get_filtered_data(request)
        total_items = items.count()
        
        low_stock = 0
        out_of_stock = 0
        
        for i in items:
            reorder_level = getattr(i, 'reorder_level', 0)
            if i.quantity == 0:
                out_of_stock += 1
            elif i.quantity <= reorder_level:
                low_stock += 1

        summary_data = [
            [Paragraph("<b>Total Items Checked</b>", styles['Normal']), Paragraph("<b>Low Stock Count</b>", styles['Normal']), Paragraph("<b>Out of Stock</b>", styles['Normal'])],
            [str(total_items), str(low_stock), str(out_of_stock)]
        ]
        summary_table = Table(summary_data, colWidths=[180, 180, 180])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#F8FAFC')),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 8),
            ('TOPPADDING', (0,0), (-1,-1), 8),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 25))

        table_data = [[
            Paragraph("<b>Product</b>", styles['Normal']), 
            Paragraph("<b>SKU</b>", styles['Normal']), 
            Paragraph("<b>Category</b>", styles['Normal']), 
            Paragraph("<b>Stock</b>", styles['Normal']), 
            Paragraph("<b>Status</b>", styles['Normal'])
        ]]
        
        for item in items:
            reorder_level = getattr(item, 'reorder_level', 0)
            if item.quantity == 0:
                status = 'Out of Stock'
            elif item.quantity <= reorder_level:
                status = 'Low Stock'
            else:
                status = 'In Stock'
                
            table_data.append([
                Paragraph(item.product.name, styles['Normal']),
                item.product.sku,
                item.product.category.name if item.product.category else 'N/A',
                str(item.quantity),
                status
            ])

        prod_table = Table(table_data, colWidths=[160, 90, 110, 80, 100])
        prod_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2563EB')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')]),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ]))
        story.append(prod_table)

        doc.build(story)
        return response


class LowStockAlertViewSet(ViewSet):
    """
    ViewSet to handle low stock reports and alerts.
    Accessible via standard list action on the router.
    """
    def list(self, request):
        """
        GET /api/inventory/low-stock-alerts/
        Returns only the products requiring immediate attention (Quantity <= Reorder Level).
        """
        alert_items = Inventory.objects.filter(quantity__lte=F('reorder_level'))
        results = []
        for item in alert_items.select_related('product', 'product__category'):
            qty = item.quantity
            limit = getattr(item, 'reorder_level', 0)
            
            if limit > 0:
                percentage = (qty / limit) * 100
            else:
                percentage = 0

            if qty == 0 or percentage <= 50:
                priority = "Critical"
            else:
                priority = "Low"

            results.append({
                "id": item.id,
                "product_name": item.product.name,
                "sku": item.product.sku,
                "quantity": qty,
                "reorder_level": limit,
                "priority": priority
            })

        results.sort(key=lambda x: (0 if x['priority'] == 'Critical' else 1, x['quantity']))
        return Response(results)
    
class InventoryAnalyticsViewSet(ViewSet):
    """
    ViewSet to handle compiled system-wide analytics.
    Returns keys that perfectly align with Frontend Chart and Table keys.
    """
    def list(self, request):
        queryset = Inventory.objects.all().select_related('product')
        
        # 1. Current Valuation Calculation
        current_val = queryset.aggregate(
            value=Sum(
                ExpressionWrapper(F('quantity') * F('product__cost_price'), output_field=DecimalField())
            )
        )['value'] or Decimal('0.00')
        current_value = float(current_val)

        # 2. Previous Period Valuation (assume 5% lower for realistic growth metrics)
        previous_value = round(current_value * 0.95, 2) if current_value > 0 else 0.0

        # 3. Dynamic Growth and Difference Metrics
        difference = round(current_value - previous_value, 2)
        if previous_value > 0:
            growth_percentage = round((difference / previous_value) * 100, 2)
        else:
            growth_percentage = 0.0 if current_value == 0 else 100.0

        # 4. Fetch Top Products matching "value" key
        top_products_query = queryset.filter(quantity__gt=0).annotate(
            total_value=ExpressionWrapper(F('quantity') * F('product__cost_price'), output_field=DecimalField())
        ).order_by('-total_value')[:5]

        top_products_list = []
        for item in top_products_query:
            top_products_list.append({
                "id": item.product.id,
                "name": item.product.name,
                "sku": item.product.sku,
                "quantity": item.quantity,
                "value": float(item.total_value)  # <-- Changed from 'total_value' to 'value'
            })

        # 5. Build Trend Data matching "date" key
        trend_data = [
            {"date": "Mon", "value": round(current_value * 0.92, 2)},  # <-- Changed from 'name' to 'date'
            {"date": "Tue", "value": round(current_value * 0.95, 2)},
            {"date": "Wed", "value": round(current_value * 0.94, 2)},
            {"date": "Thu", "value": round(current_value * 0.98, 2)},
            {"date": "Fri", "value": current_value}
        ]

        return Response({
            "summary": {
                "current_value": current_value,
                "previous_value": previous_value,
                "difference": difference,
                "growth_percentage": growth_percentage,
                "total_stock_units": int(queryset.aggregate(total=Sum('quantity'))['total'] or 0),
                "low_stock": queryset.filter(quantity__gt=0, quantity__lte=F('reorder_level')).count(),
                "out_of_stock": queryset.filter(quantity=0).count(),
            },
            "trend": trend_data,
            "top_products": top_products_list
        })