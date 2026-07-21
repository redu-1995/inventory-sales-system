import csv
import io
import openpyxl
from django.db import models, transaction
from django.http import HttpResponse
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.viewsets import ModelViewSet

from inventory.models import Inventory
from .models import Category, Product, Supplier
from .serializers import (
    CategorySerializer,
    ProductSerializer,
    SupplierSerializer,
)


class CategoryViewSet(ModelViewSet):
    """
    Handles CRUD operations for Categories.
    """
    queryset = Category.objects.all().order_by("-created_at")
    serializer_class = CategorySerializer


class SupplierViewSet(ModelViewSet):
    """
    Handles CRUD operations for Suppliers.
    """
    queryset = Supplier.objects.all().order_by("-created_at")
    serializer_class = SupplierSerializer


class ProductViewSet(ModelViewSet):
    """
    Handles CRUD operations for Products with filtering and integrated Inventory updates.
    """
    queryset = (
        Product.objects.select_related("category", "supplier", "inventory")
        .all()
        .order_by("-created_at")
    )
    serializer_class = ProductSerializer

    def get_queryset(self):
        """
        Dynamically filter products based on URL parameters:
        ?search=... &category=... &stock_status=LOW_STOCK|OUT_OF_STOCK
        """
        queryset = super().get_queryset()
        
        search_query = self.request.query_params.get("search", None)
        category_id = self.request.query_params.get("category", None)
        supplier_id = self.request.query_params.get("supplier", None)
        stock_status = self.request.query_params.get("stock_status", None)

        if search_query:
            queryset = queryset.filter(
                models.Q(name__icontains=search_query) |
                models.Q(sku__icontains=search_query) |
                models.Q(barcode__icontains=search_query)
            )

        if category_id:
            queryset = queryset.filter(category_id=category_id)

        if supplier_id:
            queryset = queryset.filter(supplier_id=supplier_id)

        if stock_status == "LOW_STOCK":
            queryset = queryset.filter(
                inventory__quantity__gt=0,
                inventory__quantity__lte=models.F("inventory__reorder_level")
            )
        elif stock_status == "OUT_OF_STOCK":
            queryset = queryset.filter(
                models.Q(inventory__quantity=0) | models.Q(inventory__isnull=True)
            )

        return queryset

    @transaction.atomic
    def perform_create(self, serializer):
        """
        Hook executed on POST requests: Saves Product and auto-creates Inventory.
        """
        # Save Product instance (this invokes ProductSerializer.create)
        product = serializer.save()
        
        # Ensure an Inventory instance exists even if not provided in serializer
        quantity = serializer.validated_data.get("inventory", {}).get("quantity", 0)
        reorder_level = serializer.validated_data.get("inventory", {}).get("reorder_level", 10)
        
        Inventory.objects.get_or_create(
            product=product,
            defaults={"quantity": quantity, "reorder_level": reorder_level}
        )


class ProductExportView(APIView):
    """
    Exports filtered inventory products into an Excel file (.xlsx).
    """
    def get(self, request):
        queryset = (
            Product.objects.select_related("category", "supplier", "inventory")
            .all()
            .order_by("-created_at")
        )

        # Apply URL Filters
        search_query = request.query_params.get("search", None)
        category_name = request.query_params.get("category", None)
        stock_status = request.query_params.get("stock_status", None)

        if search_query:
            queryset = queryset.filter(name__icontains=search_query)
        if category_name:
            queryset = queryset.filter(category__name=category_name)
        if stock_status == "LOW_STOCK":
            queryset = queryset.filter(
                inventory__quantity__gt=0,
                inventory__quantity__lte=models.F("inventory__reorder_level")
            )
        elif stock_status == "OUT_OF_STOCK":
            queryset = queryset.filter(
                models.Q(inventory__quantity=0) | models.Q(inventory__isnull=True)
            )

        # Generate openpyxl spreadsheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Inventory Status"

        # Headers
        worksheet.append(["Product Name", "SKU", "Category", "Cost Price", "Selling Price", "Quantity", "Reorder Level"])

        for product in queryset:
            inventory = getattr(product, "inventory", None)
            quantity = getattr(inventory, "quantity", 0)
            reorder_level = getattr(inventory, "reorder_level", 10)
            category_name_str = product.category.name if product.category else "Uncategorized"

            worksheet.append([
                product.name,
                product.sku,
                category_name_str,
                product.cost_price,
                product.selling_price,
                quantity,
                reorder_level
            ])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="inventory_export.xlsx"'
        workbook.save(response)
        return response


class ProductImportView(APIView):
    """
    Bulk imports products from CSV or XLSX spreadsheets.
    """
    def post(self, request):
        uploaded_file = request.FILES.get("file")
        if not uploaded_file:
            return Response({"error": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)

        filename = uploaded_file.name
        rows_data = []

        try:
            if filename.endswith(".csv"):
                data_stream = io.StringIO(uploaded_file.read().decode("utf-8"))
                reader = csv.reader(data_stream)
                next(reader, None)  # Skip table header
                for row in reader:
                    if row:
                        rows_data.append(row)

            elif filename.endswith(".xlsx"):
                workbook = openpyxl.load_workbook(uploaded_file, data_only=True)
                worksheet = workbook.active
                iterator = worksheet.iter_rows(values_only=True)
                next(iterator, None)  # Skip header
                for row in iterator:
                    if any(row):
                        rows_data.append(row)
            else:
                return Response(
                    {"error": "Unsupported file format. Please upload a .csv or .xlsx file."},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except Exception as e:
            return Response({"error": f"Failed reading spreadsheet: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

        created_count = 0
        updated_count = 0
        validation_errors = []

        with transaction.atomic():
            for index, row in enumerate(rows_data, start=2):
                try:
                    name = str(row[0]).strip()
                    sku = str(row[1]).strip()
                    category_name = str(row[2]).strip() if len(row) > 2 and row[2] else "General"
                    cost_price = float(row[3])
                    selling_price = float(row[4])
                    quantity = int(row[5])
                    reorder_level = int(row[6]) if len(row) > 6 and row[6] is not None else 10

                    # Fetch or build category reference
                    category, _ = Category.objects.get_or_create(name=category_name)

                    # Upsert Product record
                    product, created = Product.objects.update_or_create(
                        sku=sku,
                        defaults={
                            "name": name,
                            "category": category,
                            "cost_price": cost_price,
                            "selling_price": selling_price,
                            "status": True if quantity > 0 else False
                        }
                    )

                    # Update related Inventory model
                    inventory, _ = Inventory.objects.get_or_create(product=product)
                    inventory.quantity = quantity
                    inventory.reorder_level = reorder_level
                    inventory.save()

                    if created:
                        created_count += 1
                    else:
                        updated_count += 1

                except (IndexError, ValueError, TypeError) as error_info:
                    validation_errors.append(
                        f"Row {index}: Parsing error formatting text columns. Details: {str(error_info)}"
                    )
                    continue

            if validation_errors:
                transaction.set_rollback(True)
                return Response({"errors": validation_errors}, status=status.HTTP_422_UNPROCESSABLE_ENTITY)

        return Response({
            "message": "Bulk import completed successfully.",
            "created_count": created_count,
            "updated_count": updated_count
        }, status=status.HTTP_200_OK)